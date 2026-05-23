from core.ordenacao import ordenar_rotas_por_distancia
from core.pilha import Pilha
from core.fila import Fila
import csv
import os
import sys
import time
import tracemalloc
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_DIR = os.path.abspath(os.path.join(CURRENT_DIR, ".."))

if SRC_DIR not in sys.path:
    sys.path.insert(0, SRC_DIR)


if not tracemalloc.is_tracing():
    tracemalloc.start()


class SGFlowService:
    """Regras de negócio do SGFLOW."""

    def __init__(self):
        self.fila = Fila()
        self.em_atendimento = None
        self.rotas = []
        self.ordem_rotas = "crescente"
        self.metricas = []
        self._id = 1

    def _novo_id(self):
        novo_id = self._id
        self._id += 1
        return novo_id

    def _memoria_atual_mb(self):
        memoria_atual, _ = tracemalloc.get_traced_memory()
        return round(memoria_atual / (1024 * 1024), 2)

    def _registrar_metrica(self, operacao, algoritmo, inicio, volume):
        tempo_ms = (time.perf_counter() - inicio) * 1000

        self.metricas.append({
            "operacao": operacao,
            "algoritmo": algoritmo,
            "tempo_ms": round(tempo_ms, 4),
            "volume": volume,
            "memoria_mb": self._memoria_atual_mb(),
            "complexidade": "O(n²)",
        })

    def listar_metricas(self):
        return list(reversed(self.metricas))

    def limpar_metricas(self):
        self.metricas.clear()
        return True, "Dados de performance limpos com sucesso."

    def resumo_performance(self):
        total_operacoes = len(self.metricas)
        tempo_total = sum(item["tempo_ms"] for item in self.metricas)
        volume_total = sum(item["volume"] for item in self.metricas)
        memoria_total = sum(item["memoria_mb"] for item in self.metricas)

        tempo_medio = round(tempo_total / total_operacoes,
                            4) if total_operacoes else 0
        memoria_media = round(
            memoria_total / total_operacoes, 2) if total_operacoes else 0

        return {
            "total_operacoes": total_operacoes,
            "tempo_total": round(tempo_total, 4),
            "tempo_medio": tempo_medio,
            "volume_total": volume_total,
            "memoria_total": round(memoria_total, 2),
            "memoria_media_mb": memoria_media,
            "fila": len(self.fila.exibir()),
            "rotas": len(self.rotas),
            "ordem_rotas": self.ordem_rotas,
        }

    def exportar_metricas_txt(self):
        resumo = self.resumo_performance()
        linhas = [
            "SGFLOW - Relatório de Performance",
            "=================================",
            f"Total de operações: {resumo['total_operacoes']}",
            f"Tempo total: {resumo['tempo_total']} ms",
            f"Volume total: {resumo['volume_total']}",
            f"Memória média: {resumo['memoria_media_mb']} MB",
            f"Memória total registrada: {resumo['memoria_total']} MB",
            f"Ordem atual das rotas: {self.ordem_rotas}",
            "",
            "Operações:",
        ]

        if not self.metricas:
            linhas.append("Nenhuma métrica registrada.")
        else:
            for indice, item in enumerate(self.metricas, start=1):
                linhas.append(
                    f"{indice}. Operação: {item['operacao']} | "
                    f"Algoritmo/Estrutura: {item['algoritmo']} | "
                    f"Tempo: {item['tempo_ms']} ms | "
                    f"Volume: {item['volume']} | "
                    f"Memória: {item['memoria_mb']} MB | "

                )

        return "\n".join(linhas)

    def cadastrar_caminhao(self, placa, motorista, operacao, destino, distancia, cargas=None):
        inicio = time.perf_counter()
        placa = str(placa).strip().upper()
        motorista = str(motorista).strip()
        operacao = str(operacao).strip().capitalize()
        destino = str(destino).strip()
        cargas = cargas or []

        try:
            distancia = float(str(distancia).replace(",", "."))
        except ValueError:
            return False, "Distância inválida. Informe KM numérico."

        if operacao not in ["Carga", "Descarga"]:
            return False, "Operação inválida. Use Carga ou Descarga."

        if not placa or not motorista or not destino:
            return False, "Placa, motorista e destino são obrigatórios."

        if operacao == "Descarga" and len(cargas) == 0:
            return False, "Caminhão de Descarga precisa ter cargas iniciais. Preencha o campo Cargas."

        caminhao = {
            "id": self._novo_id(),
            "placa": placa,
            "motorista": motorista,
            "operacao": operacao,
            "destino": destino,
            "distancia": distancia,
            "pilha": Pilha(cargas),
            "status": "Na fila",
        }

        self.fila.enqueue(caminhao)
        self._registrar_metrica("Cadastro de caminhão",
                                "Fila FIFO", inicio, len(self.fila.exibir()))

        return True, f"Caminhão {placa} cadastrado na fila."

    def listar_fila(self):
        return self.fila.exibir()

    def selecionar(self, caminhao_id):
        caminhao_id = int(caminhao_id)

        if self.em_atendimento and self.em_atendimento["id"] == caminhao_id:
            return self.em_atendimento

        return self.fila.buscar_por_id(caminhao_id)

    def atender_proximo(self):
        inicio = time.perf_counter()

        if self.em_atendimento:
            return False, "Já existe um caminhão em atendimento. Finalize carga/descarga antes de atender o próximo."

        try:
            self.em_atendimento = self.fila.dequeue()
            self.em_atendimento["status"] = "Em atendimento"
            placa = self.em_atendimento["placa"]
            self._registrar_metrica(
                "Atender próximo da fila", "Fila FIFO", inicio, len(self.fila.exibir()))

            return True, f"Caminhão {placa} removido da fila e colocado em atendimento."
        except IndexError as erro:
            return False, str(erro)

    def adicionar_carga_carga(self, item):
        inicio = time.perf_counter()

        if not self.em_atendimento:
            return False, "Nenhum caminhão em atendimento."

        if self.em_atendimento["operacao"] != "Carga":
            return False, "Este caminhão é de Descarga."

        item = str(item).strip()

        if not item:
            return False, "Digite o item da carga."

        self.em_atendimento["pilha"].push(item)
        volume = len(self.em_atendimento["pilha"].exibir())
        self._registrar_metrica(
            "Adicionar carga", "Pilha LIFO", inicio, volume)

        return True, f"Carga '{item}' adicionada ao topo da pilha."

    def adicionar_carga_descarga(self, item):
        inicio = time.perf_counter()

        if not self.em_atendimento:
            return False, "Nenhum caminhão em atendimento."

        if self.em_atendimento["operacao"] != "Descarga":
            return False, "Esta opção é apenas para caminhão de Descarga."

        item = str(item).strip()

        if not item:
            return False, "Digite a carga recebida."

        self.em_atendimento["pilha"].push(item)
        volume = len(self.em_atendimento["pilha"].exibir())
        self._registrar_metrica(
            "Adicionar carga recebida", "Pilha LIFO", inicio, volume)

        return True, f"Carga recebida '{item}' adicionada à pilha de descarga."

    def descarregar(self):
        inicio = time.perf_counter()

        if not self.em_atendimento:
            return False, "Nenhum caminhão em atendimento."

        if self.em_atendimento["operacao"] != "Descarga":
            return False, "Este caminhão é de Carga."

        try:
            item = self.em_atendimento["pilha"].pop()
            volume = len(self.em_atendimento["pilha"].exibir())
            mensagem = f"Carga '{item}' removida do topo da pilha."
            self._registrar_metrica(
                "Descarga de item", "Pilha LIFO", inicio, volume)

            if self.em_atendimento["pilha"].esta_vazia():
                placa = self.em_atendimento["placa"]
                self._liberar_para_rota()
                mensagem += f" Caminhão {placa} vazio e liberado com sucesso."

            return True, mensagem
        except IndexError as erro:
            return False, str(erro)

    def finalizar_carga(self):
        inicio = time.perf_counter()

        if not self.em_atendimento:
            return False, "Nenhum caminhão em atendimento."

        if self.em_atendimento["operacao"] != "Carga":
            return False, "Este caminhão não é de Carga."

        if self.em_atendimento["pilha"].esta_vazia():
            return False, "Adicione pelo menos uma carga antes de liberar."

        placa = self.em_atendimento["placa"]
        self._liberar_para_rota()
        self._registrar_metrica(
            "Finalizar carga", "Pilha LIFO", inicio, len(self.rotas))

        return True, f"Caminhão {placa} carregado e liberado com sucesso."

    def _liberar_para_rota(self):
        self.em_atendimento["status"] = "Liberado para rota"
        self.rotas.append(self.em_atendimento)
        self.rotas = self._ordenar_rotas()
        self.em_atendimento = None

    def _ordenar_rotas(self):
        inicio = time.perf_counter()
        algoritmo = "Insertion Sort" if self.ordem_rotas == "decrescente" else "Bubble Sort"
        rotas_ordenadas = ordenar_rotas_por_distancia(
            self.rotas, self.ordem_rotas)
        self._registrar_metrica("Ordenação de rotas",
                                algoritmo, inicio, len(rotas_ordenadas))

        return rotas_ordenadas

    def inverter_ordem_rotas(self):
        if self.ordem_rotas == "crescente":
            self.ordem_rotas = "decrescente"
            algoritmo = "Insertion Sort"
            nome = "maior KM primeiro"
        else:
            self.ordem_rotas = "crescente"
            algoritmo = "Bubble Sort"
            nome = "menor KM primeiro"

        self.rotas = self._ordenar_rotas()

        return True, f"Ordem alterada para {nome}."

    def listar_rotas(self):
        self.rotas = ordenar_rotas_por_distancia(self.rotas, self.ordem_rotas)
        return self.rotas.copy()

    def liberar_partida(self):
        inicio = time.perf_counter()
        self.rotas = ordenar_rotas_por_distancia(self.rotas, self.ordem_rotas)

        if not self.rotas:
            return False, "Nenhum caminhão liberado para rota."

        caminhao = self.rotas.pop(0)
        self._registrar_metrica(
            "Liberar partida", "Lista de rotas", inicio, len(self.rotas))

        return True, f"Caminhão {caminhao['placa']} liberado com sucesso para partida."

    def importar_arquivo(self, caminho):
        inicio = time.perf_counter()
        caminho = Path(caminho)
        extensao = caminho.suffix.lower()

        if extensao not in [".csv", ".xlsx"]:
            return False, "Formato inválido. Somente .csv e .xlsx são aceitos."

        if extensao == ".csv":
            linhas = self._ler_csv(caminho)
        else:
            linhas = self._ler_xlsx(caminho)

        total = 0
        erros = []

        for numero_linha, linha in enumerate(linhas, start=2):
            cargas = self._parse_cargas(linha.get("cargas", ""))
            sucesso, mensagem = self.cadastrar_caminhao(
                linha.get("placa", ""),
                linha.get("motorista", ""),
                linha.get("operacao", ""),
                linha.get("destino", ""),
                linha.get("distancia", ""),
                cargas,
            )

            if sucesso:
                total += 1
            else:
                erros.append(f"Linha {numero_linha}: {mensagem}")

        self._registrar_metrica("Importação de arquivo",
                                "CSV/XLSX", inicio, total)

        if erros:
            return False, f"{total} importados. Erros: " + " | ".join(erros)

        return True, f"{total} caminhões importados para a o patio."

    def _parse_cargas(self, valor):
        return [item.strip() for item in str(valor or "").split("|") if item.strip()]

    def _ler_csv(self, caminho):
        with open(caminho, "r", encoding="utf-8-sig", newline="") as arquivo:
            leitor = csv.DictReader(arquivo)
            return [
                {str(chave).strip().lower(): valor for chave,
                 valor in linha.items()}
                for linha in leitor
            ]

    def _ler_xlsx(self, caminho):
        if load_workbook is None:
            raise RuntimeError("openpyxl não instalado.")

        planilha = load_workbook(caminho).active
        linhas = list(planilha.iter_rows(values_only=True))

        if not linhas:
            return []

        cabecalhos = [str(item).strip().lower() for item in linhas[0]]
        dados = []

        for valores in linhas[1:]:
            linha = {}

            for indice, cabecalho in enumerate(cabecalhos):
                if indice < len(valores) and valores[indice] is not None:
                    linha[cabecalho] = valores[indice]
                else:
                    linha[cabecalho] = ""

            dados.append(linha)

        return dados
